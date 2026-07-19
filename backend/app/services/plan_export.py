"""Tour-plan export — the office's printable/shareable week plan.

Both formats render the same view: header facts (customer, calendar week,
date range, assigned worker, status), then the stops grouped by assigned day
in driving order with address, ETA, closing time, service minutes, tasks and
the plan's remarks. Stops without an assigned day land in a trailing
"Unscheduled" group so nothing silently disappears from a handout.

Colors follow the design tokens (DESIGN.md): ink #16233a, muted #5b6b84,
borders #dde4ee, header fill #f6f8fb.
"""

import io
from dataclasses import dataclass
from datetime import date, datetime

from sqlalchemy.orm import Session, selectinload

from app.models.stop import Stop
from app.models.tour import Tour
from app.models.user import User

INK = "16233A"
MUTED = "5B6B84"
BORDER = "DDE4EE"
FILL = "F6F8FB"

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA = "application/pdf"


@dataclass
class PlanRow:
    day: date | None
    sequence: int | None
    market: str
    address: str
    eta: str
    closes: str
    minutes: str
    tasks: str
    remarks: str
    done: bool


def _hhmm(value: datetime | None) -> str:
    return value.strftime("%H:%M") if value else ""


def _weekday(d: date) -> str:
    return ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][d.weekday()]


def _day_label(d: date | None) -> str:
    if d is None:
        return "Unscheduled"
    return f"{_weekday(d)} {d.strftime('%d.%m.%Y')}"


def _rows(db: Session, tour: Tour) -> list[PlanRow]:
    stops = (
        db.query(Stop)
        .options(selectinload(Stop.tasks), selectinload(Stop.store))
        .filter(Stop.tour_id == tour.id)
        .all()
    )
    stops.sort(
        key=lambda s: (
            s.assigned_day is None,
            s.assigned_day or date.max,
            s.sequence if s.sequence is not None else 10_000,
            s.id,
        )
    )
    rows = []
    for s in stops:
        address = ", ".join(
            part
            for part in [
                s.effective_street,
                " ".join(p for p in [s.effective_postal_code, s.effective_city] if p),
            ]
            if part
        )
        tasks = ", ".join(t.raw_label or t.task_type for t in s.tasks)
        closing = s.effective_closing_time
        rows.append(
            PlanRow(
                day=s.assigned_day,
                sequence=s.sequence,
                market=s.customer or f"Stop {s.id}",
                address=address,
                eta=_hhmm(s.eta),
                closes=closing.strftime("%H:%M") if closing else "",
                minutes=str(s.service_minutes) if s.service_minutes else "",
                tasks=tasks,
                remarks=s.remarks_raw or "",
                done=s.completed_at is not None,
            )
        )
    return rows


def _header_lines(db: Session, tour: Tour) -> tuple[str, str]:
    worker = db.get(User, tour.assigned_user_id) if tour.assigned_user_id else None
    assigned = worker.name if worker else (tour.employee or "unassigned")
    title = f"Tour plan — {tour.customer}"
    meta = (
        f"KW {tour.calendar_week} · {tour.date_from} – {tour.date_to} · "
        f"assigned to {assigned} · status {tour.status.value} · tour #{tour.id}"
    )
    return title, meta


HEADERS = ["Day", "#", "Market", "Address", "ETA", "Closes", "Min", "Tasks", "Remarks"]


def build_xlsx(db: Session, tour: Tour) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title, meta = _header_lines(db, tour)
    wb = Workbook()
    ws = wb.active
    ws.title = f"KW {tour.calendar_week}"

    thin = Side(style="thin", color=BORDER)
    grid = Border(bottom=thin)
    ink = Font(color=INK)
    muted = Font(color=MUTED, size=10)

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color=INK)
    ws.append([meta])
    ws["A2"].font = muted
    ws.append([])

    ws.append(HEADERS)
    header_row = ws.max_row
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, size=10, color=MUTED)
        cell.fill = PatternFill("solid", fgColor=FILL)
        cell.border = Border(bottom=Side(style="thin", color=MUTED))

    last_day: object = ...
    for row in _rows(db, tour):
        day_cell = _day_label(row.day) if row.day != last_day else ""
        last_day = row.day
        market = f"✓ {row.market}" if row.done else row.market
        ws.append(
            [
                day_cell,
                row.sequence,
                market,
                row.address,
                row.eta,
                row.closes,
                row.minutes,
                row.tasks,
                row.remarks,
            ]
        )
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = ink
            cell.border = grid
            cell.alignment = Alignment(vertical="top", wrap_text=col in (8, 9))
        if day_cell:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color=INK)

    for col, width in zip(
        range(1, len(HEADERS) + 1), [16, 4, 28, 34, 7, 7, 6, 40, 30], strict=True
    ):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(db: Session, tour: Tour) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    ink = colors.HexColor(f"#{INK}")
    muted = colors.HexColor(f"#{MUTED}")
    border = colors.HexColor(f"#{BORDER}")
    fill = colors.HexColor(f"#{FILL}")

    body = ParagraphStyle(
        "body", fontName="Helvetica", fontSize=8, leading=10, textColor=ink
    )
    title_style = ParagraphStyle(
        "title", fontName="Helvetica-Bold", fontSize=14, leading=18, textColor=ink
    )
    meta_style = ParagraphStyle(
        "meta", fontName="Helvetica", fontSize=9, leading=12, textColor=muted
    )
    day_style = ParagraphStyle(
        "day", fontName="Helvetica-Bold", fontSize=10, leading=14, textColor=ink
    )

    title, meta = _header_lines(db, tour)
    story = [
        Paragraph(title, title_style),
        Paragraph(meta, meta_style),
        Spacer(0, 4 * mm),
    ]

    rows = _rows(db, tour)
    groups: dict[object, list[PlanRow]] = {}
    for row in rows:
        groups.setdefault(row.day, []).append(row)

    col_widths = [10, 52, 62, 13, 13, 11, 66, 42]
    for day, day_rows in groups.items():
        story.append(Paragraph(_day_label(day), day_style))
        data = [HEADERS[1:]]
        for r in day_rows:
            market = f"✓ {r.market}" if r.done else r.market
            data.append(
                [
                    str(r.sequence or ""),
                    Paragraph(market, body),
                    Paragraph(r.address, body),
                    r.eta,
                    r.closes,
                    r.minutes,
                    Paragraph(r.tasks, body),
                    Paragraph(r.remarks, body),
                ]
            )
        table = Table(data, colWidths=[w * mm for w in col_widths], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("TEXTCOLOR", (0, 0), (-1, 0), muted),
                    ("TEXTCOLOR", (0, 1), (-1, -1), ink),
                    ("BACKGROUND", (0, 0), (-1, 0), fill),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.4, border),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(0, 4 * mm))

    buf = io.BytesIO()
    SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    ).build(story)
    return buf.getvalue()
