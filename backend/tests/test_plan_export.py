"""Plan export (PDF/XLSX handouts).

Requires a reachable database with migrations applied; skipped otherwise.
"""

import io
from datetime import UTC, date, datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.db import SessionLocal, engine
from app.main import app
from app.models.stop import Stop
from app.models.task import Task
from app.models.tour import Tour, TourStatus


def _db_reachable() -> bool:
    try:
        with engine.connect():
            return True
    except Exception:
        return False


pytestmark = pytest.mark.skipif(not _db_reachable(), reason="database not reachable")

client = TestClient(app)


@pytest.fixture
def planned_tour():
    """A two-day plan plus one unscheduled stop."""
    db = SessionLocal()
    tour = Tour(
        customer="Export-Test Tour",
        calendar_week=31,
        date_from=date(2027, 8, 2),
        date_to=date(2027, 8, 6),
        status=TourStatus.planned,
    )
    db.add(tour)
    db.flush()
    monday, tuesday = date(2027, 8, 2), date(2027, 8, 3)
    stops = [
        Stop(
            tour_id=tour.id,
            row_index=0,
            customer="Export Markt Eins",
            claimed_street="Teststr. 1",
            claimed_postal_code="04109",
            claimed_city="Leipzig",
            assigned_day=monday,
            sequence=1,
            eta=datetime(2027, 8, 2, 8, 30, tzinfo=UTC),
            service_minutes=45,
            remarks_raw="Schlüssel beim Nachbarn",
        ),
        Stop(
            tour_id=tour.id,
            row_index=1,
            customer="Export Markt Zwei",
            assigned_day=tuesday,
            sequence=1,
            completed_at=datetime(2027, 8, 3, 9, 0, tzinfo=UTC),
        ),
        Stop(tour_id=tour.id, row_index=2, customer="Export Markt Offen"),
    ]
    db.add_all(stops)
    db.flush()
    db.add(Task(stop_id=stops[0].id, task_type="EKW", raw_label="EKW"))
    db.commit()
    tour_id = tour.id
    db.close()

    yield tour_id

    db = SessionLocal()
    db.query(Tour).filter(Tour.id == tour_id).delete()  # stops cascade
    db.commit()
    db.close()


def test_xlsx_export_contains_plan(planned_tour):
    resp = client.get(f"/tours/{planned_tour}/plan/export", params={"format": "xlsx"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert f"tour-{planned_tour}-kw31-plan.xlsx" in resp.headers["content-disposition"]

    ws = load_workbook(io.BytesIO(resp.content)).active
    text = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for row in ws.iter_rows(values_only=True)
    )
    assert "Export-Test Tour" in text
    assert "Export Markt Eins" in text
    assert "Teststr. 1, 04109 Leipzig" in text
    assert "08:30" in text
    assert "EKW" in text
    assert "✓ Export Markt Zwei" in text  # completed marker
    assert "Unscheduled" in text  # the unplanned stop is not hidden


def test_pdf_export_is_valid_pdf(planned_tour):
    resp = client.get(f"/tours/{planned_tour}/plan/export", params={"format": "pdf"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")
    assert len(resp.content) > 1000


def test_export_validation():
    assert client.get("/tours/999999/plan/export").status_code == 404
    resp = client.get("/tours/1/plan/export", params={"format": "docx"})
    assert resp.status_code == 422
